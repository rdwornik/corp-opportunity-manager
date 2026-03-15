"""Read and update Project_Codes.xlsm — find rows, add folder links.

Real column layout (sheet 'Main'):
  A: Empty
  B: Name (SE name)
  C: Account Name (= client)
  D: Opportunity Name
  E: Link to SF
  F: Booking Amount
  G: JDA Industry
  H: Stage
  I: Close Date
  J: JDA OpptyID2
  K: Date Added
  L: Next Step
  M: Folder Link (added by this tool)
"""

from __future__ import annotations

import logging
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Column indices (0-based) matching the real spreadsheet
_COL_ACCOUNT = 2  # C: Account Name (client)
_COL_OPP_NAME = 3  # D: Opportunity Name
_COL_INDUSTRY = 6  # G: JDA Industry
_COL_STAGE = 7  # H: Stage
_COL_FOLDER = 12  # M: Folder Link (1-based: 13)


@dataclass
class ProjectRow:
    """A row from the project codes spreadsheet."""

    row_number: int
    account_name: str
    opportunity_name: str
    industry: str
    stage: str
    folder_link: str


def _safe_str(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _open_workbook_readonly(excel_path: Path):
    """Open a workbook for reading, falling back to a temp copy if the file is locked."""
    try:
        return load_workbook(excel_path, read_only=True, keep_vba=True)
    except PermissionError:
        logger.info("File locked, reading from temp copy: %s", excel_path)
        tmp_dir = tempfile.mkdtemp()
        tmp_path = Path(tmp_dir) / excel_path.name
        shutil.copy2(excel_path, tmp_path)
        return load_workbook(tmp_path, read_only=True, keep_vba=True)


def find_row_by_client(excel_path: Path, client: str) -> ProjectRow | None:
    """Find the first row matching the account name (case-insensitive, substring match).

    Skips the header row (row 1).
    """
    if not excel_path.exists():
        logger.warning("Excel file not found: %s", excel_path)
        return None

    wb = _open_workbook_readonly(excel_path)
    ws = wb.active

    client_lower = client.lower()
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_value = row[_COL_ACCOUNT].value
        if cell_value and client_lower in str(cell_value).strip().lower():
            result = ProjectRow(
                row_number=row[_COL_ACCOUNT].row,
                account_name=_safe_str(row[_COL_ACCOUNT].value),
                opportunity_name=_safe_str(row[_COL_OPP_NAME].value),
                industry=_safe_str(row[_COL_INDUSTRY].value),
                stage=_safe_str(row[_COL_STAGE].value),
                folder_link=_safe_str(row[_COL_FOLDER].value) if len(row) > _COL_FOLDER else "",
            )
            wb.close()
            return result
    wb.close()
    return None


def update_folder_link(excel_path: Path, row_number: int, folder_path: str) -> bool:
    """Set the folder link (column M) for a given row.

    Returns True if the update succeeded.
    """
    if not excel_path.exists():
        logger.error("Excel file not found: %s", excel_path)
        return False

    try:
        wb = load_workbook(excel_path, keep_vba=True)
    except PermissionError:
        logger.error("Cannot write — file is locked (close Excel first): %s", excel_path)
        return False

    ws = wb.active

    # Add header if column M header is empty
    header_cell = ws.cell(row=1, column=_COL_FOLDER + 1)
    if not header_cell.value:
        header_cell.value = "Folder Link"

    ws.cell(row=row_number, column=_COL_FOLDER + 1, value=folder_path)
    wb.save(excel_path)
    wb.close()
    logger.info("Updated row %d with folder link: %s", row_number, folder_path)
    return True


def list_projects(excel_path: Path) -> list[ProjectRow]:
    """Read all project rows from the spreadsheet."""
    if not excel_path.exists():
        logger.warning("Excel file not found: %s", excel_path)
        return []

    wb = _open_workbook_readonly(excel_path)
    ws = wb.active
    rows: list[ProjectRow] = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        account = row[_COL_ACCOUNT].value
        if account is None:
            continue
        rows.append(
            ProjectRow(
                row_number=row[_COL_ACCOUNT].row,
                account_name=_safe_str(account),
                opportunity_name=_safe_str(row[_COL_OPP_NAME].value),
                industry=_safe_str(row[_COL_INDUSTRY].value),
                stage=_safe_str(row[_COL_STAGE].value),
                folder_link=_safe_str(row[_COL_FOLDER].value) if len(row) > _COL_FOLDER else "",
            )
        )

    wb.close()
    return rows
